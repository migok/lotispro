"""Export tools for AI Assistant — Excel export from query results."""

import base64
import io
import json
from datetime import datetime
from typing import Optional

from agents import RunContextWrapper, function_tool

from app.ai_assistant.context import AssistantContext


@function_tool
def export_to_excel(
    context: RunContextWrapper[AssistantContext],
    data_json: str,
    title: str,
    sheet_name: Optional[str] = "Données",
    filename: Optional[str] = None,
) -> str:
    """Exporte des données tabulaires vers un fichier Excel (.xlsx).

    Utilise ce tool quand l'utilisateur demande un export Excel ou un téléchargement de tableau.
    Le fichier est retourné en base64 — le frontend l'affichera comme bouton de téléchargement.

    Args:
        data_json: Résultat JSON de query_database (avec champ "rows") OU liste directe de dicts.
                   Ex: '{"rows": [{"nom": "A", "valeur": 10}]}'  ou  '[{"nom": "A", "valeur": 10}]'
        title: Titre descriptif du tableau (ex: "Ventes du mois de mars 2024")
        sheet_name: Nom de l'onglet Excel (défaut: "Données")
        filename: Nom du fichier sans extension (défaut: généré depuis le titre + date)

    Returns:
        JSON avec success, excel_base64, filename, row_count, column_count

    Exemples:
        export_to_excel(data_json=query_result, title="Ventes Mars 2024")
        export_to_excel(data_json='[{"commercial": "Jean", "ventes": 5}]', title="Performance")
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return json.dumps({
            "success": False,
            "error": "openpyxl non installé. Ajoutez 'openpyxl' aux dépendances.",
        }, ensure_ascii=False)

    try:
        # Parse input — supports both query_database result format and direct list
        parsed = json.loads(data_json)
        if isinstance(parsed, dict):
            rows = parsed.get("rows", [])
        elif isinstance(parsed, list):
            rows = parsed
        else:
            rows = []

        if not rows:
            return json.dumps({
                "success": False,
                "error": "Aucune donnée à exporter (rows vide)",
            }, ensure_ascii=False)

        columns = list(rows[0].keys())

        # ── Workbook setup ────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (sheet_name or "Données")[:31]  # Excel sheet name max 31 chars

        # LotisPro color palette
        BRASS = "D4973A"
        DARK_BG = "0F1624"
        SUBHEADER_BG = "1E2840"
        ALT_ROW = "F0F2F8"
        HEADER_TEXT = "FFFFFF"

        # ── Row 1: Title ──────────────────────────────────────────────────
        last_col = get_column_letter(len(columns))
        ws.merge_cells(f"A1:{last_col}1")
        title_cell = ws["A1"]
        title_cell.value = title
        title_cell.font = Font(name="Calibri", bold=True, size=14, color=HEADER_TEXT)
        title_cell.fill = PatternFill("solid", fgColor=DARK_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

        # ── Row 2: Export date ────────────────────────────────────────────
        ws.merge_cells(f"A2:{last_col}2")
        date_cell = ws["A2"]
        date_cell.value = f"Exporté le {datetime.utcnow().strftime('%d/%m/%Y à %H:%M')} UTC"
        date_cell.font = Font(name="Calibri", italic=True, size=9, color="AAAAAA")
        date_cell.fill = PatternFill("solid", fgColor=SUBHEADER_BG)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        # ── Row 3: Spacer ─────────────────────────────────────────────────
        ws.row_dimensions[3].height = 6

        # ── Row 4: Column headers ─────────────────────────────────────────
        header_border = Border(bottom=Side(style="thin", color="C8A040"))
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = col_name.replace("_", " ").title()
            cell.font = Font(name="Calibri", bold=True, size=11, color=HEADER_TEXT)
            cell.fill = PatternFill("solid", fgColor=BRASS)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border
        ws.row_dimensions[4].height = 26

        # ── Rows 5+: Data ─────────────────────────────────────────────────
        for row_idx, row_data in enumerate(rows, start=5):
            bg_color = ALT_ROW if row_idx % 2 == 0 else "FFFFFF"
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(col_name)
                # Convert whole floats to int for cleaner display
                if isinstance(value, float) and value == int(value):
                    value = int(value)
                cell.value = value
                cell.font = Font(name="Calibri", size=10)
                cell.fill = PatternFill("solid", fgColor=bg_color)
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row_idx].height = 20

        # ── Auto column widths ────────────────────────────────────────────
        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            header_len = len(col_name.replace("_", " ").title())
            data_len = max(
                (len(str(row.get(col_name) or "")) for row in rows),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max(header_len, data_len) + 4, 45)

        # Freeze panes below header
        ws.freeze_panes = "A5"

        # ── Serialize to base64 ───────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        excel_base64 = base64.b64encode(buffer.read()).decode("utf-8")

        # Build filename
        safe_title = "".join(
            c for c in title if c.isalnum() or c in " _-"
        ).strip().replace(" ", "_")[:40]
        date_str = datetime.utcnow().strftime("%Y%m%d")
        final_filename = filename or f"{safe_title}_{date_str}"

        return json.dumps({
            "success": True,
            "excel_base64": excel_base64,
            "filename": final_filename,
            "row_count": len(rows),
            "column_count": len(columns),
            "columns": columns,
            "message": (
                f"Fichier Excel prêt: {final_filename}.xlsx "
                f"({len(rows)} lignes, {len(columns)} colonnes)"
            ),
        }, ensure_ascii=False)

    except json.JSONDecodeError as e:
        return json.dumps({
            "success": False,
            "error": f"Données JSON invalides: {str(e)}",
        }, ensure_ascii=False)
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": f"Erreur export Excel: {str(e)}",
        }, ensure_ascii=False)
